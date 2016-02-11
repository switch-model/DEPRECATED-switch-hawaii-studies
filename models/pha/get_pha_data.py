import os, sys
path_to_core = os.path.abspath(os.path.join(os.path.dirname(__file__), 'switch-hawaii-core'))
sys.path.append(path_to_core)
from scenario_data import write_tab_file
import numpy as np

try:
    import openpyxl
except ImportError:
    print "This script requires the openpyxl module to access the data in Microsoft Excel files."
    print "Please execute 'sudo pip install openpyxl' or 'pip install openpyxl' (Windows)."
    raise

# get_scenario_data.py should already have been run, creating a folder with standard inputs
inputs_dir = "inputs_tiny"

base_year = 2015
n_scenarios = 117
n_digits = 4 # len(str(n_scenarios-1))  # how many leading zeros to use for scenario names

pha_dir = os.path.join(inputs_dir, "pha")

if not os.path.exists(pha_dir):
    os.makedirs(pha_dir)

# read standard data
with open(os.path.join(inputs_dir, "fuel_supply_curves.tab")) as f:
    standard_fuel_costs = [r.split("\t") for r in f.read().splitlines()]

headers = standard_fuel_costs.pop(0)
period_ind = headers.index("period")
periods = sorted(set(float(r[period_ind]) for r in standard_fuel_costs))

# date when random forecast begins
fuel_base_date = 2015+5.0/12    # 6/1/15

# fossil fuel base prices, $/mmbtu (oil includes delivery to Hawaii)
oil_base_price = 111.67/6.115  # from Ulupono spreadsheet, for lsfo-diesel blend delivered to HI
gas_base_price = 2.889   # from Ulupono spreadsheet, for Henry Hub gas

# factors to calculate various fuel prices from base prices; 
# each is a tuple of (oil multiplier, gas multiplier, constant adder).
# factors come from "HECO fuel cost forecasts.xlsx" unless otherwise noted
price_factors = {}
price_factors["LSFO-Diesel-Blend", "base"] = (1.0, 0.0, 0.0)
price_factors["LSFO", "base"] = (13.35/13.90, 0.0, 0.0)
price_factors["Diesel", "base"] = (14.72/13.90, 0.0, 0.0)
price_factors["Biodiesel", "base"] = (14.72/13.90, 0.0, 9.217) # diesel price + delta
price_factors["LNG", "bulk"] = (0.0, 1.2, 6.0) # from Ulupono spreadsheet
price_factors["LNG", "container"] = (0.0, 1.0, 17.59)

# build an empirical version of the joint distribution of price changes in 
# oil (lsfo/diesel blend) and LNG, based on behavior in 2000-2015.
# also factor out inflation of about 0.19% per month during this period
# opening the workbook is the slowest part of the whole script
# print "loading 'Fuel Costs Monte Carlo-3.xlsx'"
wb = openpyxl.load_workbook("Fuel Costs Monte Carlo-3.xlsx", data_only=True)
ws = wb["DBEDT-HEI-BBG Fuel Jan'00-pres"]
# print "reading data from 'Fuel Costs Monte Carlo-3.xlsx'"
oil_historical_prices = np.array([r[0].value for r in ws["K3:K182"]])
gas_historical_prices = np.array([r[0].value for r in ws["N3:N182"]])
oil_historical_multipliers = (1 - 0.0019) * oil_historical_prices[1:]/oil_historical_prices[:-1]
gas_historical_multipliers = (1 - 0.0019) * gas_historical_prices[1:]/gas_historical_prices[:-1]
# print "finished reading from 'Fuel Costs Monte Carlo-3.xlsx'"

# list of all months from the forecast start date till the start of the last period
months = np.arange(fuel_base_date, periods[-1]+1.0/12, 1.0/12)
# indices of the months that are closest to the start of each period.
period_starts = np.array([np.argmin(np.abs(months - p)) for p in periods])

# define a random walk through the future months, drawing from historical variations
sample_draws = np.random.randint(0, len(oil_historical_multipliers), size=(len(months), n_scenarios))
oil_multipliers = np.cumprod(oil_historical_multipliers[sample_draws], axis=0)[period_starts, :]
gas_multipliers = np.cumprod(gas_historical_multipliers[sample_draws], axis=0)[period_starts, :]

# np.set_printoptions(suppress=True, linewidth=150)
# print "oil_multipliers"
# print oil_multipliers
# print "gas_multipliers"
# print gas_multipliers

oil_prices = oil_base_price * oil_multipliers
gas_prices = gas_base_price * gas_multipliers

# print "oil_prices"
# print oil_prices
# print "gas_prices"
# print gas_prices

fuel_prices = {
    (fuel, tier): om * oil_prices + gm * gas_prices + a 
        for (fuel, tier), (om, gm, a) in price_factors.iteritems()
}

# find indices of index columns
index_i = [headers.index(c) for c in ['regional_fuel_market', 'period', 'tier']]
# find indices of columns that should be omitted from the .dat file
# (these shouldn't be in the .tab file either, but they don't cause trouble there)
drop_i = [headers.index(c) for c in ['fuel']]
# helper function to sort a row of data from the .tab file into a suitable order for the .dat file
# (index columns must be shifted to the start of the row)
sort_cols = lambda r: (
    [r[i] for i in index_i] + [c for (i, c) in enumerate(r) if i not in index_i + drop_i]
)
# translate column names to the correct form for the model 
# (runph doesn't translate like our normal data loading functions)
col_name_dict = dict(zip(
        ["unit_cost", "max_avail_at_cost", "fixed_cost"],
        ["rfm_supply_tier_cost", "rfm_supply_tier_limit", "rfm_supply_tier_fixed_cost"]
    ))
translate_cols = lambda r: [col_name_dict.get(c, c) for c in r]

# write the data files for each scenario, using scenario-specific prices where available
for s in range(n_scenarios):
    fuel_data = []
    for row in standard_fuel_costs:
        # columns: 
        # regional_fuel_market, fuel, period, tier, unit_cost, max_avail_at_cost, fixed_cost
        r = list(row)   # copy the existing list before changing, or convert tuple to list
        if (r[1], r[3]) in fuel_prices:
            r[4] = str(fuel_prices[r[1], r[3]][periods.index(float(r[2])), s])
        fuel_data.append(r)
    # write_tab_file(
    #     "fuel_supply_curves_{s}.tab".format(s=str(s).zfill(n_digits)),
    #     headers, fuel_data,
    #     dict(inputs_dir=pha_dir, inputs_subdir="")
    # )
    with open(
        os.path.join(pha_dir, "fuel_supply_curves_{s}.dat".format(s=str(s).zfill(n_digits))),
        "w"
    ) as f:
        # omit headers for index cols
        f.write('param: ' + '\t'.join(translate_cols(sort_cols(headers)[3:])) + ' :=\n')
        for r in fuel_data:
            f.write('\t'.join(sort_cols(r)) + '\n')
        f.write(';\n')
    
