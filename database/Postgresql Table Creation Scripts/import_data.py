#!/usr/bin/env python

# This will eventually be a good place to put all the code to create the main database
# for Switch-Hawaii. For now it just imports a few miscellaneous tables that were not
# imported by other code.
# Started by Matthias Fripp on 2015-07-27

import sys, csv, datetime, os
from textwrap import dedent
import numpy as np
import pandas as pd
import sqlalchemy

sys.path.append(os.path.join('..', 'build_database'))
import shared_tables, tracking_pv

switch_db = 'switch'
pghost = 'redr.eng.hawaii.edu'
db_engine = sqlalchemy.create_engine('postgresql://' + pghost + '/' + switch_db)

try:
    import openpyxl
except ImportError:
    print "This script requires the openpyxl module to access the data in Microsoft Excel files."
    print "Please execute 'sudo pip install openpyxl' or 'pip install openpyxl' (Windows)."
    raise

try:
    import psycopg2
except ImportError:
    print "This script requires the psycopg2 module to access the postgresql database."
    print "Please execute 'sudo pip install psycopg2' or 'pip install psycopg2' (Windows)."
    raise
try:
    # note: the connection gets created when the module loads and never gets closed 
    # (until presumably python exits)
    con = psycopg2.connect(database=switch_db, host=pghost, sslmode='require')
    # note: we don't autocommit because it makes executemany() very slow; 
    # instead we call con.commit() after each query
    con.autocommit = False


except psycopg2.OperationalError:
    print dedent("""
        ############################################################################################
        Error while connecting to {db} database on postgresql server {server}.
        Please ensure that your user name on the local system is the same as your postgresql user 
        name or there is a local PGUSER environment variable set with your postgresql user name.
        There should also be a line like "*:*:{db}:<username>:<password>" in ~/.pgpass or 
        %APPDATA%\postgresql\pgpass.conf (Windows). On Unix systems, .pgpass should be chmod 0600.
        See http://www.postgresql.org/docs/9.3/static/libpq-pgpass.html for more details.
        ############################################################################################
        """.format(db=switch_db, server=pghost))
    raise


def main():
    # run all the import scripts (or at least the ones we want)
    # ev_adoption()
    # fuel_costs()
    # energy_source_properties()
    # fuel_costs_no_biofuel()
    # rps_timeseries()
    # system_load()
    # generator_info()
    # onshore_wind()
    # offshore_wind()
    tracking_pv.tracking_pv()
    tracking_pv.distributed_pv()
    shared_tables.calculate_interconnect_costs()

def execute(query, arguments=None):
    args = [dedent(query)]
    if arguments is not None:
        args.append(arguments)
    cur = con.cursor()
    cur.execute(*args)
    con.commit()
    return cur

def executemany(query, arguments=None):
    args = [dedent(query)]
    if arguments is not None:
        args.append(arguments)
    cur = con.cursor()
    cur.executemany(*args)
    con.commit()

open_workbooks = {}

def get_workbook(xlsx_file):
    if xlsx_file not in open_workbooks:
        # load the file, ignore formula text
        open_workbooks[xlsx_file] = openpyxl.load_workbook(xlsx_file, data_only=True)
    return open_workbooks[xlsx_file]

def get_table_from_xlsx(xlsx_file, named_range, transpose=False):
    wb = get_workbook(xlsx_file)
    full_range = wb.get_named_range(named_range)
    # note: named range should be a simple rectangular region; 
    # if it contains more than one region we ignore all but the first
    d1 = full_range.destinations[0]
    ws = d1[0]
    region = d1[1]
    data = list(tuple(c.value for c in r) for r in ws[region])
    if transpose:
        data = zip(*data)
    head = data.pop(0)  # take out the header row
    data = zip(*data)   # switch from row to column orientation
    # make a dictionary, with one column for each element of header row
    return dict(zip(head, data))
    
def get_named_region(xlsx_file, named_range):
    # get a single rectangular region from the specified file in the source data directory
    wb = get_workbook(xlsx_file)
    full_range = wb.get_named_range(named_range)
    if full_range is None:
        raise ValueError(
            'Range "{}" not found in workbook "{}".'.format(named_range, xlsx_file))
    if len(full_range.destinations) > 1:
        raise ValueError(
            'Range "{}" in workbook "{}" contains more than one region.'.format(named_range, xlsx_file))
    ws, region = full_range.destinations[0]
    return ws[region]

def data_frame_from_xlsx(xlsx_file, named_range):
    region = get_named_region(xlsx_file, named_range)
    return pd.DataFrame([cell.value for cell in row] for row in region)

def get_named_cell_from_xlsx(xlsx_file, named_range):
    region = get_named_region(xlsx_file, named_range)
    if not isinstance(region, openpyxl.cell.cell.Cell):
        raise ValueError(
            'Range "{}" in workbook "{}" does not refer to an individual cell.'.format(
                named_range, xlsx_file))
    return region.value

#########################
# rps timeseries (reusing version from before the server crashed)
def rps_timeseries():
    execute("""
        delete from study_date where time_sample='rps';
        delete from study_hour where time_sample='rps';    
    """)

    with open('timeseries_rps.tab','r') as f:
        for r in csv.DictReader(f, delimiter='\t'):
            dt = str(r["TIMESERIES"])[2:8]
            execute("""
                INSERT INTO study_date 
                    (time_sample, period, study_date, 
                    month_of_year, date, 
                    hours_in_sample,
                    ts_num_tps, ts_duration_of_tp, ts_scale_to_period)
                VALUES
                    (%s, %s, %s,
                    %s, %s,
                    %s,
                    %s, %s, %s);
            """,
                ('rps', r["ts_period"], r["TIMESERIES"], 
                int(dt[2:4]), "20"+dt[0:2]+"-"+dt[2:4]+"-"+dt[4:6],
                float(r["ts_duration_of_tp"])*float(r["ts_scale_to_period"]),
                r["ts_num_tps"], r["ts_duration_of_tp"], r["ts_scale_to_period"])
            )
    
    with open('timepoints_rps.tab','r') as f:
        for r in csv.DictReader(f, delimiter='\t'):
            i += 1
            sys.stdout.write('row: {}\r'.format(i))
            sys.stdout.flush()
            t = r["timestamp"][5:]
            dt = str(r["timeseries"])[2:8]
            execute("""
                INSERT INTO study_hour 
                    (time_sample, study_date, study_hour,
                    hour_of_day, date_time)
                VALUES
                    (%s, %s, %s,
                    %s, 
                    cast(%s as timestamp with time zone));
            """,
                ('rps', r["timeseries"], r["timepoint_id"], 
                int(t[6:8]), 
                "20" + dt[0:2] + '-' + t[:5] + ' ' + t[6:] + ':00-10')
            )

    execute("""
        delete from study_periods where time_sample='rps';
        insert into study_periods
            select distinct time_sample, period from study_date
            where time_sample='rps'
            order by 2;

        -- mini RPS study (even months, even hours) for reasonable results in relatively short time

        drop table if exists tdoublemonths;
        create temporary table tdoublemonths
            (month_of_year smallint primary key, days_in_month smallint);
        insert into tdoublemonths values 
          (1, 59), (2, 59), (3, 61), (4, 61), (5, 61), (6, 61),
          (7, 62), (8, 62), (9, 61), (10, 61), (11, 61), (12, 61);

        delete from study_date where time_sample='rps_mini';
        insert into study_date 
            (period, study_date, month_of_year, date, 
            hours_in_sample, time_sample, ts_num_tps, ts_duration_of_tp, ts_scale_to_period)
            select period, study_date, d.month_of_year, date,
                0.0 as hours_in_sample, 
                'rps_mini' as time_sample,
                12 as ts_num_tps, 2.0 as ts_duration_of_tp,
                case when ts_scale_to_period < 100 then 2*%(years_per_period)s 
                    else (days_in_month-2)*%(years_per_period)s end as ts_scale_to_period
            from study_date d join tdoublemonths m using (month_of_year)
            where time_sample = 'rps' and mod(month_of_year, 2)=0
            order by 1, 3, 5 desc, 4;

        delete from study_hour where time_sample='rps_mini';
        insert into study_hour (study_date, study_hour, hour_of_day, date_time, time_sample)
          select h.study_date, study_hour, hour_of_day, date_time, 
            'rps_mini' as time_sample
          from study_hour h join study_date d on (d.time_sample='rps_mini' and d.study_date=h.study_date)
          where h.time_sample = 'rps' and mod(hour_of_day, 2)=0 
          order by period, month_of_year, hours_in_sample desc, hour_of_day;

        delete from study_periods where time_sample='rps_mini';
        insert into study_periods
            select distinct time_sample, period from study_date
            where time_sample='rps_mini'
            order by 2;


    """, dict(years_per_period=8))


    print "Created rps and rps_mini time samples."


#########################
# ev adoption
def ev_adoption():
    # identify pairs of (ev_scen_id, HECO scenario name):
    ev_adoption_scenarios=(
        (0, 'Business as Usual'), # very low
        (1, 'No Burning Desire'), # low
        (2, 'Blazing a Bold Frontier'), # high
        (3, 'Stuck in the Middle'), # medium, a.k.a. 'Moved by Passion'
        (4, 'Full Adoption'), # 100% by 2045
    )
    # get the EV adoption curves from an Excel workbook
    # uses logistic curves fitted to HECO IRP 2013 Appendix E-10, p. E-113,
    # as well as VMT data from DBEDT Economic Databook
    # and vehicle registration rates from DBEDT monthly energy spreadsheet
    ev_adoption_curves = get_table_from_xlsx("EV projections.xlsx", named_range='ev_data')

    # create the ev_adoption table
    execute("""
        DROP TABLE IF EXISTS ev_adoption;
        CREATE TABLE ev_adoption (
            load_zone varchar(40),
            year int,
            ev_scenario varchar(40),
            ev_share float,
            ice_miles_per_gallon float,
            ev_miles_per_kwh float,
            ev_extra_cost_per_vehicle_year float,
            n_all_vehicles float,
            vmt_per_vehicle float
        );
    """)

    # insert data into the ev_adoption table
    n_rows = len(ev_adoption_curves['Year'])
    for (ev_scen_id, ev_scenario_name) in ev_adoption_scenarios:
        executemany(
            "INSERT INTO ev_adoption VALUES ({})".format(','.join(["%s"]*9)),
            zip(
                ['Oahu']*n_rows, ev_adoption_curves['Year'], [ev_scenario_name]*n_rows, 
                ev_adoption_curves[ev_scenario_name], # % adoption
                ev_adoption_curves["ICE miles per gallon"],
                ev_adoption_curves["EV miles per kWh"],
                ev_adoption_curves["EV extra cost per vehicle per year"],
                ev_adoption_curves["number of vehicles"],
                ev_adoption_curves["VMT per vehicle"],
            )
        )

    print "Created ev_adoption table."
    
    # see /Users/matthias/Dropbox/Research/shared/Paritosh/M.S Thesis Paritosh/Data Used In Thesis/calculate BAU charging.ipynb
    
    # create the ev_hourly_charge_profile table (simple business-as-usual charging profile,
    # given as hourly weights)
    execute("""
        DROP TABLE IF EXISTS ev_hourly_charge_profile;
        CREATE TABLE ev_hourly_charge_profile (
            hour_of_day smallint, 
            charge_weight float
        );
    """)
    with open('ev_hourly_charge_profile.tsv') as f:
        profile = [r.split("\t") for r in f.read().splitlines()][1:] # skip headers
        
    executemany(
        "INSERT INTO ev_hourly_charge_profile (hour_of_day, charge_weight) VALUES (%s, %s);",
        profile
    )
    print "Created ev_hourly_charge_profile table."
    


def fuel_costs():
    # TODO: add base_year to fuel_costs table, and fuel_cost import scripts,
    # and use it for inflation adjustments in scenario_data.py

    # create the fuel_costs table if needed
    execute("""
        CREATE TABLE IF NOT EXISTS fuel_costs (
            load_zone varchar(40),
            year int,
            fuel_type varchar(30),
            price_mmbtu float,
            fixed_cost float,
            max_avail_at_cost float,
            fuel_scen_id varchar(40),
            tier varchar(20),
        );
        ALTER TABLE fuel_costs OWNER TO admin;
    """)

    # TODO: add fixed_cost and max_avail_at_cost for EIA-based forecasts
    
    # Oahu fuel price forecasts, derived from EIA
    import_fuel_costs("../../../data/EIA-based fuel cost forecasts/HECO fuel cost forecasts.xlsx", 'EIA_ref')
    import_fuel_costs("../../../data/EIA-based fuel cost forecasts/HECO fuel cost forecasts_low.xlsx", 'EIA_low')
    import_fuel_costs("../../../data/EIA-based fuel cost forecasts/HECO fuel cost forecasts_high.xlsx", 'EIA_high')
    import_fuel_costs("../../../data/EIA-based fuel cost forecasts/HECO fuel cost forecasts_LNG_pegged_to_oil.xlsx", 'EIA_lng_oil_peg')
    import_fuel_costs("../../../data/EIA-based fuel cost forecasts/HECO fuel cost forecasts_high_LNG_pegged_to_oil.xlsx", 'EIA_high_lng_oil_peg')

    # Oahu hedged fuel costs
    import_hedged_fuel_costs("../../../data/EIA-based fuel cost forecasts/hedged fuel prices.xlsx", "hedged")

def import_fuel_costs(file, fuel_scen_id):
    
    # get the forecasts from an Excel workbook 
    # Based on various sources, cited in the workbook, extended to 2050
    fuel_forecast = get_table_from_xlsx(file, named_range='Adjusted_EIA_Forecast', transpose=True)

    # remove any existing records
    execute("""
        DELETE FROM fuel_costs WHERE fuel_scen_id=%s;
    """, (fuel_scen_id,))

    # take out the list of years, so the dictionary just has one entry for each fuel
    years = fuel_forecast.pop('Year')

    # insert data into the fuel_costs table
    n_rows = len(years)
    for f in fuel_forecast:
        ft=f.split(", ")
        fuel = ft[0]
        tier = ft[1] if len(ft) >= 2 else 'base'
        executemany("""
            INSERT INTO fuel_costs (load_zone, year, fuel_type, price_mmbtu, fuel_scen_id, tier) 
            VALUES (%s, %s, %s, %s, %s, %s)""",
            zip(['Oahu']*n_rows, 
                years,
                [fuel]*n_rows,
                fuel_forecast[f], 
                [fuel_scen_id]*n_rows,
                [tier]*n_rows
            )
        )

    print "Added EIA-derived forecast (fuel_scen_id={}) to fuel_costs table.".format(fuel_scen_id)


def import_hedged_fuel_costs(file, fuel_scen_id):
    
    prices = data_frame_from_xlsx(file, named_range='fuel_prices')
    prices = prices.set_index(0)
    prices.index.name = 'year'
    prices = prices.T.set_index(['fuel_type', 'tier']).T.astype(float)
    # switch to one row per value, and assign a name to the value
    prices = pd.DataFrame({'price_mmbtu': prices.stack(['fuel_type', 'tier'])})
    prices['load_zone'] = 'Oahu'
    prices['fuel_scen_id'] = fuel_scen_id

    tiers = data_frame_from_xlsx(file, named_range='tier_properties')
    # Transpose, set row and column labels, and convert to floating point (converting None to NaN)
    tiers = tiers.set_index(0).T.set_index(['fuel_type', 'tier']).astype(float)

    # join the two together (have to have drop and restore the year index to make this work)
    prices = prices.reset_index('year').join(tiers).set_index('year', append=True)
    # could follow with prices = prices.reorder_levels(['year', 'fuel_type', 'tier']) if wanted

    # remove any existing records
    execute("DELETE FROM fuel_costs WHERE fuel_scen_id=%s;", (fuel_scen_id,))
    
    prices.to_sql('fuel_costs', db_engine, if_exists='append')
    
    print "Added hedged prices (fuel_scen_id={}) to fuel_costs table.".format(fuel_scen_id)

#########################
# Fuel properties, maintained manually in the Excel forecast workbook 
def energy_source_properties():
    properties = get_table_from_xlsx(
        "../../../Reference Docs/HECO Plans/HECO fuel cost forecasts.xlsx", 
        named_range='Fuel_Properties'
    )

    # create the fuel_properties table if needed
    execute("""
        CREATE TABLE IF NOT EXISTS energy_source_properties (
            energy_source VARCHAR(30) PRIMARY KEY,      -- name of the fuel
            fuel_rank DECIMAL(4, 2),           -- usually 1-5, but may be decimal, e.g., 1.5
            rps_eligible SMALLINT,             -- 0 or 1
            co2_intensity FLOAT                -- tCO2 per MMBtu
        );
    """)

    # create a temporary table to hold the data before aggregating by fuel type
    execute("""
        DROP TABLE IF EXISTS t_energy_source_properties;
        CREATE TEMPORARY TABLE t_energy_source_properties (LIKE energy_source_properties);
    """)

    # insert data into the energy_source_properties table
    executemany("""
        INSERT INTO t_energy_source_properties (energy_source, fuel_rank, rps_eligible, co2_intensity) 
        VALUES (%s, %s, %s, %s)""",
        zip(
            [f.split(', ')[0] for f in properties['Fuel']], 
            properties['Rank'],
            properties['RPS Eligible'],
            [i/1000.0 for i in properties['kg CO2 per MMbtu']], 
        )
    )

    # move the data into the main energy_source_properties table
    execute("""
        DELETE FROM energy_source_properties;
        INSERT INTO energy_source_properties SELECT DISTINCT * FROM t_energy_source_properties;
        DROP TABLE t_energy_source_properties;
    """)

    print "Created energy_source_properties table."

def fuel_costs_no_biofuel():
    """Create no-biofuel fuel cost scenarios"""
    execute("""
        DELETE FROM fuel_costs WHERE fuel_scen_id LIKE 'EIA_%_no_biofuel';
        INSERT INTO fuel_costs
        SELECT load_zone, year, c.fuel_type, price_mmbtu, 
            fuel_scen_id || '_no_biofuel' as fuel_scen_id,
            tier
        FROM fuel_costs c JOIN energy_source_properties p ON c.fuel_type = p.energy_source
        WHERE rps_eligible = 0 AND fuel_scen_id LIKE 'EIA_%';
    """)

def onshore_wind():
    """Import old onshore wind data into newer tables."""
    # TODO: write code to create these records directly from OWITS data and GIS files
    # and also store location and interconnect distance
    execute("""
        delete from cap_factor 
            where project_id in 
                (select project_id from project where technology = 'OnshoreWind');
        delete from project where technology='OnshoreWind';
        insert into project 
            (load_zone, technology, site, orientation, max_capacity, connect_distance_km)
            select load_zone, technology, concat('OnWind_', site), 
                orientation, max_capacity, connect_length_km
            from max_capacity_pre_2016_06_21 
                left join connect_cost_pre_2016_06_21 using (load_zone, technology, site, orientation)
                where technology='OnshoreWind';
        insert into cap_factor (project_id, date_time, cap_factor)
            select project_id, date_time, cap_factor
                from cap_factor_pre_2016_06_21 c
                    join project p on 
                        (p.load_zone=c.load_zone and p.technology=c.technology and 
                        p.site = concat('OnWind_', c.site) and p.orientation=c.orientation)
                    where c.technology='OnshoreWind';
    """)
    
def offshore_wind():
    """Import approximate capacity factor for offshore wind farms. This is calculated
    as the average of three proposed offshore sites to get approximately the right 
    amount for diversified offshore wind. (It might be better just to model them as
    three separate projects.)
    
    Note: The 2016 PSIP used hourly output possibly for 2014, from an existing 
    wind farm on the Big Island with a capacity factor of 42%. We don't use this 
    because it's the wrong profile for offshore Oahu, and especially because it
    has inconsistent timing with our other weather and load data so it would 
    create an artificial appearance of diversity (strong winds when Oahu actually 
    has windless/sunless days).
    """
    # approximate locations for the centers of three proposed wind farms
    # were found on 2016-04-07 by inspecting the 
    # "Atlantic and Pacific OCS Wind Planning Areas and Wind Energy Areas" 
    # shapefile from http://www.boem.gov/Renewable-Energy-GIS-Data/
    # (http://www.boem.gov/uploadedFiles/BOEM/Renewable_Energy_Program/Mapping_and_Data/Wind_Planning_Areas.zip)
    locs = np.array([[21.656, -158.572], [21.096, -157.987], [20.969, -157.799]])
    cells = np.array(list(execute("select i, j, lat, lon from cell where grid_id = 'E';")))
    cell_lat_lon = cells[:,-2:]
    # this makes one row for each site, one col for each cell, showing distance in degrees**2
    dist2 = ((locs[:,np.newaxis,:] - cell_lat_lon[np.newaxis,:,:])**2).sum(axis=2)
    match_cells = dist2.argmin(axis=1)
    turbine_cells = cells[match_cells]
    
    # normalized power curve for generic offshore wind turbine from http://www.nrel.gov/docs/fy14osti/61714.pdf p. 5,
    # with operating range extended to 30 m/s like Repower 6 M shown on p. 4.
    power_curve = np.array(zip(
        range(32), 
        [0] * 4 + [0.0281, 0.074, 0.1373, 0.2266, 0.3443, 0.4908, 0.6623, 0.815, 0.9179, 0.9798]
        + [1] * 17 + [0]
    ))
        
    hourly = pd.DataFrame(
        list(execute(
            """
                select grid_id, i, j, complete_time_stamp as date_time, s100 
                    from hourly_average 
                    where grid_id='E' and (i, j) in %(cells)s;""", 
            {"cells": tuple(tuple(c for c in r) for r in turbine_cells[:,:2].astype(int))}
        )),
        columns=['grid_id', 'i', 'j', 'date_time', 's100'],
    )
    hourly.set_index(['date_time', 'grid_id', 'i', 'j'], inplace=True)
    
    hourly['p100'] = np.interp(hourly['s100'].values, power_curve[:,0], power_curve[:,1])
    hourly = hourly.unstack(level=['grid_id', 'i', 'j'])
    # use mean of three sites as hourly output; also derate same as we do for land sites (from IRP 2013)
    hourly['power'] = hourly['p100'].mean(axis=1)*0.8747

    # delete any old OffshoreWind records from cap_factor
    execute("""
        delete from cap_factor 
            where project_id in 
                (select project_id from project where load_zone = 'Oahu' and technology = 'OffshoreWind');
    """)
  
    # add the new project to the project table
    execute("""
        delete from project where technology = 'OffshoreWind' and load_zone = 'Oahu';
        insert into project 
            (load_zone, technology, site, orientation, max_capacity)
            values ('Oahu', 'OffshoreWind', 'OffWind', 'na', 800);
    """)
    # retrieve the project_id for the new project
    project_id = execute("""
        select project_id from project where load_zone = 'Oahu' and technology = 'OffshoreWind';
    """).next()[0]
    
    # put the power data into cap_factor
    out_data = zip([project_id]*len(hourly), hourly.index.astype(datetime.datetime), hourly['power'].values)
    
    # TODO: consider removing and restoring indexes before this
    executemany("""
        insert into cap_factor (project_id, date_time, cap_factor)
        values (%s, %s, %s);
    """, out_data)
    
    # note: we don't add latitude, longitude or interconnect_id (and cost) because we don't 
    # have project-specific connection costs for them. So they will automatically use
    # the generic connection cost from generator_info (assigned later). 
    # That happens to be zero in this case since the connection cost is included in the overnight cost.

def generator_info():
    """Read data from 'PSIP 2016 generator data.xlsx and it in generator_info and generator_costs_by_year.

    This spreadsheet is based on HECO's 2016-04-01 PSIP assumptions, shown in the report and sent in a
    separate spreadsheet on 2016-05-05."""
    
    base_year = get_named_cell_from_xlsx('PSIP 2016 generator data.xlsx', 'o_m_base_year')

    gen_info = data_frame_from_xlsx('PSIP 2016 generator data.xlsx', 'technology_info')
    # set column headers and row indexes (index in the dataframe become index in the table)
    gen_info = gen_info.T.set_index(0).T.set_index('technology')
    gen_info.rename(
        columns={
            'fixed_o_m_per_kw_year': 'fixed_o_m', 
            'variable_o_m_per_mwh': 'variable_o_m',
            'full_load_heat_rate': 'heat_rate'
        }, inplace=True)
    # convert from cost per MWh to cost per kWh
    gen_info['variable_o_m'] *= 0.001   
    # convert unit_size = NA or 0 to NaN
    gen_info['unit_size'] = gen_info['unit_size'].where(
        (gen_info['unit_size'] != "NA") & (gen_info['unit_size'] != 0)
    )
    # report base_year for inflation calculations later
    gen_info['base_year'] = base_year
    
    # convert all columns except fuel to numeric values (some of these were imported 
    # as objects due to invalid values, which have now been removed)
    # gen_info.convert_objects() does this nicely, but is deprecated.
    for c in gen_info.columns:
        if c != 'fuel':
            gen_info[c] = pd.to_numeric(gen_info[c])
    # store data
    gen_info.to_sql('generator_info', db_engine, if_exists='replace')
    
    # load gen capital cost info
    gen_costs = data_frame_from_xlsx('PSIP 2016 generator data.xlsx', 'technology_costs')
    inflation_rate = get_named_cell_from_xlsx('PSIP 2016 generator data.xlsx', 'inflation_rate')
    
    gen_costs = gen_costs.T.set_index(0).T
    gen_costs.columns.name = 'technology'
    # drop info rows (now 0-6)
    gen_costs = gen_costs.iloc[7:]
    # rename the first column
    gen_costs = gen_costs.rename(columns={gen_costs.columns[0]: 'year'})
    # set year index
    gen_costs = gen_costs.set_index('year')
    # convert N/A to nan
    gen_costs = gen_costs.where(gen_costs != ' N/A ').astype(float)

    # select only the technologies that are in gen_info
    gen_costs = gen_costs[gen_info.index]

    # convert to real dollars in base year
    gen_costs = gen_costs.multiply(
        (1 + inflation_rate) ** (base_year - gen_costs.index.values), 
        axis='rows'
    )

    # switch to stacked orientation (one row per year, tech), but keep as a DataFrame
    gen_costs = pd.DataFrame({'capital_cost_per_kw': gen_costs.stack()})
    # record the base year to allow adjustment to other years later
    gen_costs['base_year'] = base_year
    gen_costs.to_sql('generator_costs_by_year', db_engine, if_exists='replace')

def system_load():
    # TODO: extend to other load zones by adding more rows to 'PSIP 2016 generator data.xlsx!sales_forecast'
    
    # get historical peak and average loads
    hist = pd.read_sql(
        sql="""
            SELECT 
                load_zone, EXTRACT(year FROM date_time) as year_hist, 
                MAX(system_load) as peak_hist, AVG(system_load) as avg_hist
            FROM system_load 
            GROUP BY 1, 2;
        """,
        con=db_engine
    )
    # forecast peak and energy
    fore = data_frame_from_xlsx('PSIP 2016 generator data.xlsx', 'sales_forecast')
    fore = fore.T.set_index(0).T
    fore = fore.rename(columns={'year': 'year_fore'})
    # calculate scale factors
    sls = pd.merge(hist, fore, on='load_zone')
    sls['load_scen_id'] = 'PSIP_2016_04'
    sls['peak_fore'] = sls['underlying forecast (MW)'] + sls['energy efficiency (MW)']
    sls['avg_fore'] = (sls['underlying forecast (GWh)'] + sls['energy efficiency (GWh)'])/8.76
    sls['scale'] = (sls['peak_fore'] - sls['avg_fore']) / (sls['peak_hist'] - sls['avg_hist'])
    sls['offset'] = sls['peak_fore'] - sls['scale'] * sls['peak_hist']

    # put into standard order, drop unneeded columns, convert to the right types for the database
    db_columns = [
        'load_zone', 'load_scen_id', 'year_hist', 'year_fore', 
        'peak_hist', 'peak_fore', 'avg_hist', 'avg_fore', 'scale', 'offset'
    ]
    system_load_scale = pd.DataFrame()
    for c in db_columns:
        if c in ['load_zone', 'load_scen_id']:
            system_load_scale[c] = sls[c].astype(str)
        elif c in ['year_hist', 'year_fore']:
            system_load_scale[c] = sls[c].astype(int)
        else:
            system_load_scale[c] = sls[c].astype(float)
    system_load_scale.set_index(db_columns[:4], inplace=True)
    # store data
    execute("""delete from system_load_scale where load_scen_id='PSIP_2016_04';""")
    system_load_scale.to_sql('system_load_scale', db_engine, if_exists='append')

def interconnect():
    # also see database/build_database/shared_tables.py for code to fill in
    # project.interconnect_id, project.connect_distance_km and project.connect_cost_per_mw
    # based on this table
    # note: we could eventually add interconnect-specific connection costs here, 
    # to be used instead of generic project interconnection costs; in that case
    # the code in shared_tables.calculate_interconnect_costs() would also need 
    # to be updated
    execute("""
        DROP TABLE IF EXISTS interconnect;
        CREATE TABLE interconnect (
            interconnect_id integer PRIMARY KEY NOT NULL,
            county text,
            latitude float,
            longitude float
        );
        ALTER TABLE interconnect OWNER TO admin;
        -- At some point interconnect was filled in with the equivalent of the 
        -- following command. The original code is missing, but these appear to be 
        -- the population-weighted centers of each county.
        INSERT INTO interconnect (interconnect_id, county, latitude, longitude) VALUES 
            (1, 'Honolulu', 21.372464, -157.913673),
            (2, 'Hawaii', 19.672837, -155.421895),
            (3, 'Maui', 20.863747, -156.493816),
            (4, 'Kauai', 22.021022, -159.442112),
            (5, 'Kalawao', 21.188495, -156.979972);
    """)


if __name__ == "__main__":
    main()
    con.close()